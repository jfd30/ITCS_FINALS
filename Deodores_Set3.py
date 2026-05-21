import os
import tkinter as tk
from tkinter import messagebox, ttk
import openpyxl as op


def initialize_excel():
    filename = "ordersDB.xlsx"
    if not os.path.exists(filename):
        wb = op.Workbook()
        ws = wb.active
        ws.title = "Orders"

        headers = [
            "Order ID",
            "Customer Name",
            "Product",
            "Quantity",
            "Price",
            "Total",
        ]
        ws.append(headers)

        initial_data = [
            [1, "Juan Dela Cruz", "Burger", 2, 75, 150],
            [2, "Maria Santos", "Fries", 3, 50, 150],
            [3, "Carlo Reyes", "Pizza", 1, 350, 350],
            [4, "Angela Lopez", "Milktea", 4, 120, 480],
            [5, "Kevin Ramos", "Spaghetti", 2, 95, 190],
        ]

        for row in initial_data:
            ws.append(row)

        wb.save(filename)


initialize_excel()

selected_item_id = None
selected_order_id = None

def load_data():
    for row in table.get_children():
        table.delete(row)

    wb = op.load_workbook("ordersDB.xlsx")
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  
            table.insert("", "end", values=row)


def on_row_select(event):
    global selected_item_id, selected_order_id

    selected_items = table.selection()
    if not selected_items:
        return

    selected_item_id = selected_items[0]
    values = table.item(selected_item_id, "values")

    if values:
        selected_order_id = values[0] 

        cname_entry.delete(0, tk.END)
        product_entry.delete(0, tk.END)
        qty_entry.delete(0, tk.END)
        price_entry.delete(0, tk.END)

        cname_entry.insert(0, values[1])
        product_entry.insert(0, values[2])
        qty_entry.insert(0, values[3])
        price_entry.insert(0, values[4])


def validate_inputs():
    if (
        not cname_entry.get().strip()
        or not product_entry.get().strip()
        or not qty_entry.get().strip()
        or not price_entry.get().strip()
    ):
        messagebox.showerror("Error", "All fields must be filled out!")
        return False

    try:
        qty = int(qty_entry.get())
        if qty <= 0:
            raise ValueError
    except ValueError:
        messagebox.showerror(
            "Error", "Quantity must be a positive whole number!"
        )
        return False

    try:
        price = float(price_entry.get())
        if price <= 0:
            raise ValueError
    except ValueError:
        messagebox.showerror("Error", "Price must be a positive number!")
        return False

    return True


def update_record():
    global selected_order_id

    if not selected_order_id:
        messagebox.showwarning(
            "Warning", "Please select a record from the table to update."
        )
        return

    if not validate_inputs():
        return

    wb = op.load_workbook("ordersDB.xlsx")
    ws = wb.active

    cname = cname_entry.get().strip()
    product = product_entry.get().strip()
    qty = int(qty_entry.get())
    price = float(price_entry.get())
    total = qty * price

    row_found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == int(selected_order_id):
            row[1].value = cname
            row[2].value = product
            row[3].value = qty
            row[4].value = price
            row[5].value = total
            row_found = True
            break

    if row_found:
        wb.save("ordersDB.xlsx")
        load_data()  
        clear_entries()
        messagebox.showinfo("Success", "Record updated successfully!")
    else:
        messagebox.showerror("Error", "Could not find matching record in database.")


def delete_record():
    global selected_order_id

    if not selected_order_id:
        messagebox.showwarning(
            "Warning", "Please select a record from the table to delete."
        )
        return

    confirm = messagebox.askyesno(
        "Confirm Delete", "Are you sure you want to delete this record?"
    )
    if not confirm:
        return

    wb = op.load_workbook("ordersDB.xlsx")
    ws = wb.active

    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == int(selected_order_id):
            ws.delete_rows(r_idx, 1)
            break

    wb.save("ordersDB.xlsx")
    load_data()  
    clear_entries()
    messagebox.showinfo("Success", "Record deleted successfully!")


def clear_entries():
    global selected_order_id, selected_item_id
    cname_entry.delete(0, tk.END)
    product_entry.delete(0, tk.END)
    qty_entry.delete(0, tk.END)
    price_entry.delete(0, tk.END)
    selected_order_id = None
    selected_item_id = None


window = tk.Tk()
window.title("Simple Ordering System")
window.configure(bg="lightblue")

title = tk.Label(
    window,
    text="Simple Ordering System",
    font=("Times New Roman", 14, "bold"),
    bg="lightblue",
)
title.grid(row=0, column=0, columnspan=6, pady=5)

genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=7, padx=10, pady=10)

cname_entry = tk.Entry(genframe, font=("Poppins", 12))
cname_entry.grid(row=2, column=1, columnspan=2, padx=10, pady=(10, 0))

cname_label = tk.Label(
    genframe, text="Customer Name", font=("Poppins", 10, "italic"), bg="lightblue"
)
cname_label.grid(row=3, column=1, columnspan=2)

product_entry = tk.Entry(genframe, font=("Poppins", 12))
product_entry.grid(row=2, column=3, columnspan=2, padx=10, pady=(10, 0))

product_label = tk.Label(
    genframe, text="Product", font=("Poppins", 10, "italic"), bg="lightblue"
)
product_label.grid(row=3, column=3, columnspan=2)

qty_entry = tk.Entry(genframe, font=("Poppins", 12))
qty_entry.grid(row=4, column=1, columnspan=2, padx=10, pady=(10, 0))

qty_label = tk.Label(
    genframe, text="Quantity", font=("Poppins", 10, "italic"), bg="lightblue"
)
qty_label.grid(row=5, column=1, columnspan=2)

price_entry = tk.Entry(genframe, font=("Poppins", 12))
price_entry.grid(row=4, column=3, columnspan=2, padx=10, pady=(10, 0))

price_label = tk.Label(
    genframe, text="Price", font=("Poppins", 10, "italic"), bg="lightblue"
)
price_label.grid(row=5, column=3, columnspan=2)

submit_btn = tk.Button(
    window, text="Submit", font=("Poppins", 12, "bold"), bg="lightpink"
)
submit_btn.grid(row=6, column=1, pady=(10, 20))

update_btn = tk.Button(
    window,
    text="Update",
    font=("Poppins", 12, "bold"),
    bg="lightgreen",
    command=update_record,
)
update_btn.grid(row=6, column=2, pady=(10, 20))

delete_btn = tk.Button(
    window,
    text="Delete",
    bg="red",
    fg="white",
    font=("Poppins", 12, "bold"),
    command=delete_record,
)
delete_btn.grid(row=6, column=3, pady=(10, 20))

table = ttk.Treeview(
    window,
    columns=("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"),
    show="headings",
)

for headings in (
    "Order ID",
    "Customer Name",
    "Product",
    "Quantity",
    "Price",
    "Total",
):
    table.heading(headings, text=headings)
    table.column(headings, width=100, anchor="center")

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)

table.bind("<<TreeviewSelect>>", on_row_select)

load_data()

window.mainloop()